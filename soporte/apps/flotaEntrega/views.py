from django.views.generic import ListView,TemplateView
from django.shortcuts import render
from tablib import Dataset
from apps.flotaEntrega.resource import FlotaEntregaResource
from apps.flotaEntrega.models import FlotaEntrega
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin

# Create your views here.
class FlotaEntregaList(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = ('flotaEntrega.view_flotaentrega')
    model = FlotaEntrega
    template_name = 'flotaEntrega/index.html'

def insert(request):
    if request.method == 'POST':  
        try: 
            flota_resource = FlotaEntregaResource()  
            dataset = Dataset()  
            nuevas_entregas = request.FILES['xlsFile']  
            dataset.load(nuevas_entregas.read())  
            result = flota_resource.import_data(dataset, dry_run=True)  
            print(result.has_errors())
            if not result.has_errors():  
                flota_resource.import_data(dataset, dry_run=False) # Actually import now   
        except Exception as e:
            print("error",e)
            
    return render(request, 'flotaEntrega/insert.html')  
    
class FlotaEntregaExport(PermissionRequiredMixin,TemplateView):
    permission_required = ('flotaEntrega.change_flotaentrega')
    def get(self,request,*args,**kwarg):
        entregas = FlotaEntrega.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'VIN'
        ws['B1'] = 'PLACA'
        ws['C1'] = 'FECHA VENTA'
        ws['D1'] = 'FECHA ENTREGA'

        rows = 2
        for entrega in entregas:
            ws.cell(row= rows, column= 1).value = entrega.vin
            ws.cell(row= rows, column= 2).value = entrega.placa
            ws.cell(row= rows, column= 3).value = entrega.fecha_venta
            ws.cell(row= rows, column= 4).value = entrega.fecha_entrega   
            rows += 1
            
        fileName = "ReporteEntregas.xlsx"
        response = HttpResponse(content_type= "application/ms-excel")
        content = "attachment; filename ={0}".format(fileName)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class FlotaEntregaReportView(TemplateView):
        template_name = 'flotaEntrega/report.html'
    
class FlotaEntregaReportDownload(TemplateView):
    template_name = 'flotaEntrega/report.html'
    def get(self,request,*args,**kwarg):
        startDate = request.GET.get('startDate')
        endDate = request.GET.get('endDate')
        entregas = FlotaEntrega.objects.filter(fecha_entrega__range =[startDate, endDate])
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'VIN'
        ws['B1'] = 'PLACA'
        ws['C1'] = 'FECHA VENTA'
        ws['D1'] = 'FECHA ENTREGA'

        rows = 2
        for entrega in entregas:
            ws.cell(row= rows, column= 1).value = entrega.vin
            ws.cell(row= rows, column= 2).value = entrega.placa
            ws.cell(row= rows, column= 3).value = entrega.fecha_venta
            ws.cell(row= rows, column= 4).value = entrega.fecha_entrega   
            rows += 1
            
        fileName = "ReporteEntregas.xlsx"
        response = HttpResponse(content_type= "application/ms-excel")
        content = "attachment; filename ={0}".format(fileName)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
class FlotaEntregaChart(TemplateView):
    template_name = 'flotaEntrega/chart.html'
    
    def get_context_data(self,**kwarg):
        #startDate = request.GET.get('startDate')
        context = super().get_context_data(**kwarg)
        context["qs"] =  FlotaEntrega.objects.all()[:6]#filter(fecha_entrega__range = startDate)
        return context