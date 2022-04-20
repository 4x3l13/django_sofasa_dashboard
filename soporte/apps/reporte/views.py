from django.http import Http404
from rest_framework import viewsets
from .models import Reporte
from .serializers import ReporteSerializer
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.views.generic import ListView,CreateView,UpdateView,DeleteView,TemplateView
from .models import Reporte
from .forms import InsertForm
from .services import get_data_api,post_data_api,delete_data_api,put_data_api
from django.shortcuts import redirect
from django.db import connection
from openpyxl import Workbook
from django.http.response import HttpResponse

class ReporteViewSet(viewsets.ModelViewSet):
    queryset = Reporte.objects.all()
    serializer_class = ReporteSerializer
    
class ReporteIndex(LoginRequiredMixin,PermissionRequiredMixin,ListView):
    permission_required = ('reporte.view_reporte')
    template_name = 'reporte/index.html'
    
    def get_queryset(self):
        return get_data_api('http://127.0.0.1:8000/reporte/api/')
    
class ReporteDownload(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    permission_required = ('reporte.download_reporte')
    template_name = 'reporte/download.html'
    
    def get_context_data(self,**kwarg):
        dict = get_data_api('http://127.0.0.1:8000/reporte/api/')
        context = super().get_context_data(**kwarg)
        context["object_list"] = dict
        return context
    
    def post(self,request, *args, **kwargs):
        dict = request.POST
        parameters = ''
        parametersList = []
        for row in dict:
            if row not in ('csrfmiddlewaretoken','report','sp'):
                parameters += '%s,'
                parametersList.append(dict[row])
                print(row)
        cursor = connection.cursor()
        cursor.execute(dict['sp']+ parameters, parametersList)

        columns = [i[0] for i in cursor.description]
        rows = cursor.fetchall()
        cursor.close()
        
        wb = Workbook()
        ws = wb.active
        rowcel = 1
        colcel = 1
        for col in columns:         
            ws.cell(row= rowcel, column= colcel).value = col
            colcel += 1

        rowcel += 1
        for row in rows:
            colcel = 1
            for col in columns:
                ws.cell(row= rowcel, column= colcel).value = row[colcel-1]
                colcel += 1
            rowcel += 1
            
        fileName = "Reporte.xlsx"
        response = HttpResponse(content_type= "application/ms-excel")
        content = "attachment; filename ={0}".format(fileName)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
class ReporteInsert(LoginRequiredMixin,PermissionRequiredMixin,CreateView):
    permission_required = ('reporte.add_reporte')
    form_class = InsertForm
    template_name = "reporte/insert.html"
    success_url = "/reporte"
    
    def post(self,request, *args, **kwargs):
        dict = request.POST
        post_data_api('http://127.0.0.1:8000/reporte/api/',dict)
        return redirect('reporte_index')
    
class ReporteEdit(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    permission_required = ('reporte.change_reporte')
    form_class = InsertForm
    template_name = "reporte/update.html"
    success_url = "/reporte"
    
    def get_context_data(self,pk, **kwarg):
        data = get_data_api('http://127.0.0.1:8000/reporte/api/' + str(pk) + '/')
        context = super().get_context_data(**kwarg)
        context["object_list"] = data
        return context
    
    def post(self,request, *args, **kwargs):
        dict = request.POST
        pk = request.POST.get('id')
        put_data_api('http://127.0.0.1:8000/reporte/api/'+ str(pk) + '/',dict)
        return redirect('reporte_index')
    
class ReporteDelete(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    permission_required = ('reporte.delete_reporte')
    template_name = "reporte/delete.html"
    
    def get_context_data(self,pk, **kwarg):
        data = get_data_api('http://127.0.0.1:8000/reporte/api/' + str(pk) + '/')
        context = super().get_context_data(**kwarg)
        context["object_list"] = data
        return context
    
    def post(self,request, *args, **kwargs):
        pk = request.POST.get('id')
        delete_data_api('http://127.0.0.1:8000/reporte/api/'+ str(pk) + '/')
        return redirect('reporte_index')