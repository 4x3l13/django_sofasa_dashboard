from django.contrib.auth.decorators import login_required, permission_required 
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import TemplateView
from django.db import connection
from openpyxl import Workbook
from django.http.response import HttpResponse
from .services import get_data_api

# Create your views here.
@login_required
@permission_required('encuesta.view_encuesta')
def encuestaIndex(request):
    cursor = connection.cursor()
    cursor.execute('exec SP_ENCUESTA_GETPERIOD')
    columns = [i[0] for i in cursor.description]
    rows = cursor.fetchall()
    cursor.close()
    return render(request,'encuesta/index.html',{'rows': rows, 'columns':columns})

# class EncuestaSearch(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
#     permission_required = ('encuesta.search_encuesta')
#     template_name = 'encuesta/search.html'
    
#     def post(self,request, *args, **kwargs):
#         vin = self.request.POST.get('VIN')
#         context = self.get_context_data(**kwargs)
#         cursor = connection.cursor()
#         cursor.execute('EXEC SP_ENCUESTA_GETINFORMATIONVIN %s',[vin])
#         columns = [i[0] for i in cursor.description]
#         context["rows"] = cursor.fetchall()
#         context["columns"] = columns
#         cursor.close()
#         return render(request,self.template_name,context)
    
class EncuestaSearchAPI(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    permission_required = ('encuesta.search_encuesta')
    template_name = 'encuesta/searchAPI.html'
    
    def post(self,request, *args, **kwargs):
        vin = self.request.POST.get('VIN')
        url = 'http://127.0.0.1:8000/myapi/apiEncuestaBuscarVin/?format=json&vin=' + vin
        context = self.get_context_data(**kwargs)
        context["data"] = get_data_api(url)
        return render(request,self.template_name,context)

class EncuestaChart(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    permission_required = ('encuesta.graph_encuesta')
    template_name = 'encuesta/chart.html'
    
    def post(self,request, *args, **kwargs):
        #print(request.POST)
        period = self.request.POST.get('period')
        chartjs = self.request.POST.get('chart')
        context = self.get_context_data(**kwargs)
        context["chartjs"] = chartjs
        cursor = connection.cursor()
        cursor.execute('EXEC SP_ENCUESTA_GETCHART %s',[period])
        context["rows"] = cursor.fetchall()
        cursor.close()
        context["period"] = period
        return render(request,self.template_name,context)

class EncuestaDownload(LoginRequiredMixin,PermissionRequiredMixin,TemplateView):
    
    permission_required = ('encuesta.download_encuesta')
    template_name = 'encuesta/index.html'
    
    def post(self,request, *args, **kwargs):
        #print(request.POST)
        period = self.request.POST.get('period')
        cursor = connection.cursor()
        cursor.execute('EXEC SP_ENCUESTA_GETINFORMATIONPERIOD %s',[period])

        columns = [i[0] for i in cursor.description]
        rows = cursor.fetchall()
        cursor.close()
        
        wb = Workbook()
        ws = wb.active
        rowcel = 1
        colcel = 1
        for col in columns:         
            ws.cell(row= rowcel, column= colcel).value = col
            colcel += 1

        rowcel += 1
        for row in rows:
            colcel = 1
            for col in columns:
                ws.cell(row= rowcel, column= colcel).value = row[colcel-1]
                colcel += 1
            rowcel += 1
            
        fileName = "ReporteEncuestas" + period + ".xlsx"
        response = HttpResponse(content_type= "application/ms-excel")
        content = "attachment; filename ={0}".format(fileName)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

    # def get(self,request,*args,**kwarg):
    #     period = request.GET.get('period')
    #     print(period)
    #     cursor = connection.cursor()
    #     cursor.execute('select * from flotaEntrega_flotaentrega where fecha_venta = %s',[period])
        
    #     entregas = cursor.fetchall()
    #     wb = Workbook()
    #     ws = wb.active
    #     ws['A1'] = 'VIN'
    #     ws['B1'] = 'PLACA'
    #     ws['C1'] = 'FECHA VENTA'
    #     ws['D1'] = 'FECHA ENTREGA'

    #     rows = 2
    #     for entrega in entregas:
    #         ws.cell(row= rows, column= 1).value = entrega.vin
    #         ws.cell(row= rows, column= 2).value = entrega.placa
    #         ws.cell(row= rows, column= 3).value = entrega.fecha_venta
    #         ws.cell(row= rows, column= 4).value = entrega.fecha_entrega   
    #         rows += 1
            
    #     fileName = "ReporteEntregas.xlsx"
    #     response = HttpResponse(content_type= "application/ms-excel")
    #     content = "attachment; filename ={0}".format(fileName)
    #     response['Content-Disposition'] = content
    #     wb.save(response)
    #     return response
    
    # def get_context_data(self,**kwarg):
    #     period = self.request.GET.get('period')
    #     chartjs = self.request.GET.get('chart')
    #     context = super().get_context_data(**kwarg)
    #     context["chartjs"] = chartjs
    #     cursor = connection.cursor()
    #     cursor.execute('select count(vin) vin,fecha_venta from flotaEntrega_flotaentrega group by fecha_venta')
    #     context["qs"] = cursor.fetchall()
    #     cursor.close()
    #     return context